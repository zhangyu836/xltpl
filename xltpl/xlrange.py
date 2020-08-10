# -*- coding: utf-8 -*-

from __future__ import print_function
from copy import copy

from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.cell import coordinate_to_tuple

from .utils import TreeProperty
from .utils import parse_range_tag, parse_cell_tag
from .pos import SheetPos, HRangePos, VRangePos


class XlRange(CellRange):

    node_map = TreeProperty('node_map')
    cell_map = TreeProperty('cell_map')
    range_map = TreeProperty('range_map')
    cell_tag_map = TreeProperty('cell_tag_map')

    def __init__(self, range_string=None, min_col=None, min_row=None,
                 max_col=None, max_row=None, title=None, index_base=1):
        name = ''
        _replacement = ''
        if range_string:
            parts = range_string.split('|')
            range_string = parts[0]
            if len(parts) > 1:
                name = parts[1]
            if len(parts) > 2:
                _replacement = parts[2]
        CellRange.__init__(self, range_string, min_col, min_row, max_col, max_row, title)
        self._children = []
        self._children_split = []
        self.index_base = index_base
        self.offset = offset = 1 - index_base
        self.min_rowx = self.min_row - offset
        self.max_rowx = self.max_row - offset
        self.min_colx = self.min_col - offset
        self.max_colx = self.max_col - offset
        self.name = name
        self._replacement = _replacement
        self.range_tag = None
        self.copy_count = 0
        self.copy_no = 0

    def __str__(self):
        fmt = '%s -> %s'
        return fmt % (self.__class__.__name__, self.rkey)

    @property
    def rkey(self):
        return '%s, %d%s' % (self.coord, self.depth, self.copy_no*'*')

    @property
    def depth(self):
        if not hasattr(self, 'parent') or self.parent is None or self.parent is self:
            return 0
        else:
            return self.parent.depth + 1

    @property
    def replacement(self):
        if self._replacement:
            return self.get_range(self._replacement)

    def print(self):
        print('\t' * self.depth, self)
        for child in self._children:
            child.print()

    def print_split(self):
        print('\t' * self.depth, self)
        for child in self._children_split:
            child.print_split()

    def print_range_map(self):
        for k,v in self.range_map.items():
            print(k, v)

    def equal_col_span(self, other):
        return self.min_col == other.min_col and self.max_col == other.max_col

    def equal_row_span(self, other):
        return self.min_row == other.min_row and self.max_row == other.max_row

    def row_span_intersect(self, other):
        return (self.min_row < other.min_row and other.min_row < self.max_row < other.max_row) \
               or (other.min_row < self.min_row and self.min_row < other.max_row < self.max_row)

    def add_child(self, new_range):
        for old_range in list(self._children):
            if old_range == new_range:
                print(old_range, "equal to", new_range)
                return
            elif old_range > new_range:
                old_range.add_child(new_range)
                return
            elif old_range < new_range:
                new_range.add_child(old_range)
                self._children.remove(old_range)
        new_range.parent = self
        for index,old_range in enumerate(self._children):
            if not new_range.isdisjoint(old_range):
                raise Exception('intersect', new_range, old_range)
            if new_range.row_span_intersect(old_range):
                raise Exception('row span intersect', new_range, old_range)
            if new_range.min_row < old_range.min_row or \
                    (new_range.equal_col_span(old_range) and new_range.min_col < old_range.min_col):
                self._children.insert(index, new_range)
                return
        self._children.append(new_range)

    def split(self):
        if not self._children:
            return
        min_row = self.min_row
        while self._children:
            center_range = self._children.pop(0)
            if min_row < center_range.min_row:
                up_range = HorizRange(min_row=min_row, max_row=center_range.min_row - 1,
                                      min_col=self.min_col, max_col=self.max_col, index_base=self.index_base)
                up_range.parent = self
                self._children_split.append(up_range)

            mid_range = HorizRange(min_row=center_range.min_row, max_row=center_range.max_row,
                                   min_col=self.min_col, max_col=self.max_col, index_base=self.index_base)
            mid_range.parent = self
            self._children_split.append(mid_range)
            min_col = self.min_col
            while True:
                if min_col < center_range.min_col:
                    left_range = VertRange(min_row=center_range.min_row, max_row=center_range.max_row,
                                           min_col=min_col, max_col=center_range.min_col - 1, index_base=self.index_base)
                    left_range.parent = mid_range
                    mid_range._children_split.append(left_range)
                    for child in list(self._children):
                        if child < left_range:
                            left_range.add_child(child)
                            self._children.remove(child)
                    left_range.split()

                center_range.parent = mid_range
                mid_range._children_split.append(center_range)
                center_range.split()

                min_col = center_range.max_col + 1
                if self._children:
                    next = self._children[0]
                    if next.equal_row_span(center_range):
                        center_range = self._children.pop(0)
                    else:
                        break
                else:
                    break

            if self.max_col >= min_col:
                right_range = VertRange(min_row=center_range.min_row, max_row=center_range.max_row,
                                        min_col=min_col, max_col=self.max_col, index_base=self.index_base)
                right_range.parent = mid_range
                mid_range._children_split.append(right_range)
                for child in list(self._children):
                    if child < right_range:
                        right_range.add_child(child)
                        self._children.remove(child)
                right_range.split()
            min_row = center_range.max_row + 1

        if min_row <= self.max_row:
            down_range = HorizRange(min_row=min_row, max_row=self.max_row,
                                    min_col=self.min_col, max_col=self.max_col, index_base=self.index_base)
            down_range.parent = self
            self._children_split.append(down_range)

    def copy(self, parent):
        new_range = object.__new__(self.__class__)
        new_range.__init__(min_col=self.min_col, min_row=self.min_row,
                           max_col=self.max_col, max_row=self.max_row, index_base=self.index_base)
        self.copy_count += 1
        new_range.copy_no = self.copy_count
        new_range.parent = parent
        new_range.range_tag = copy(self.range_tag)
        new_range._replacement = self._replacement
        return new_range

    def copy_subtree(self, parent):
        _self = self.copy(parent)
        for child in self._children_split:
            _child = child.copy_subtree(_self)
            _self._children_split.append(_child)
        return _self

    def copy_replacement(self, parent):
        _replacement = self.replacement.copy_subtree(parent)
        if self.range_tag:
            if not _replacement.range_tag:
                _replacement.range_tag = RangeTag()
            _replacement.range_tag.merge(self.range_tag)
        return _replacement

    def to_tpl(self):
        if self._children_split:
            tags = []
            for child in self._children_split:
                if child.replacement:
                    index = self._children_split.index(child)
                    child = child.copy_replacement(self)
                    self._children_split[index] = child
                range_tag = child.range_tag
                if range_tag and range_tag.beforerange:
                    tags.append(range_tag.beforerange)
                tags.append(child.to_tpl())
                if range_tag and range_tag.afterrange:
                    tags.append(range_tag.afterrange)
            range_tpl = '\n'.join(tags)
        else:
            range_tpl = range_to_tpl(self)
        return range_tpl

    def get_pos(self, wtsheet, pos_parent=None):
        pos_parent = self.pos_cls(wtsheet, self, pos_parent)
        for child in self._children_split:
            p = child.get_pos(wtsheet, pos_parent)
        return pos_parent

    def get_range(self, range_name):
        return self.range_map.get(range_name)

def range_to_tpl(xlrange):
    rkey = xlrange.rkey
    tags = []
    for rowx in range(xlrange.min_rowx, xlrange.max_rowx + 1):
        row_tag = "{%% row '%s' %%}" % rkey
        tag = xlrange.cell_tag_map.get((rowx, xlrange.min_colx))
        if tag and tag.beforerow:
            tags.append(tag.beforerow)
        tags.append(row_tag)
        for colx in range(xlrange.min_colx, xlrange.max_colx + 1):
            cell = xlrange.cell_map[(rowx, colx)]
            tag = xlrange.cell_tag_map.get((rowx, colx))
            if tag and tag.beforecell:
                tags.append(tag.beforecell)
            tags.append(cell.to_tag())
            if tag and tag.aftercell:
                tags.append(tag.aftercell)
    return '\n'.join(tags)


class SheetRange(XlRange):
    pos_cls = SheetPos

    def __init__(self, *args, **kwargs):
        XlRange.__init__(self, *args, **kwargs)
        self._node_map = {}
        self._cell_map = {}
        self._cell_tag_map = {}
        self._range_map = {}

    def add_cell(self, cell):
        rowx = cell.rowx
        colx = cell.colx
        self._cell_map[(rowx, colx)] = cell
        cell.parent = self

    def add_cell_tag(self, cell_tag_map, rowx, colx):
        self._cell_tag_map[(rowx, colx)] = CellTag(cell_tag_map)

    def add_range(self, range_str, range_tag):
        new_range = VertRange(range_str, index_base=self.index_base)
        self.add_child(new_range)
        key = new_range.name or new_range.coord
        self.range_map[key] = new_range
        if range_tag:
            new_range.range_tag = RangeTag(range_tag)

    def parse_tag(self, lines, rowx, colx):
        lines = lines.split(';;')
        for line in lines:
            range_str, range_tag = parse_range_tag(line)
            if range_str:
                self.add_range(range_str, range_tag)
            else:
                cell_coord, cell_tag = parse_cell_tag(line)
                if cell_coord:
                    cell_rowx, cell_colx = coordinate_to_tuple(cell_coord)
                else:
                    cell_rowx = rowx
                    cell_colx = colx
                if cell_tag:
                    self.add_cell_tag(cell_tag, cell_rowx, cell_colx)
                else:
                    pass
                    #print(rowx, colx, 'no tag')
                    #print(line)

    def get_root(self, root_name):
        root = self.get_range(root_name)
        return root or self

    def to_tpl(self, root_name='main'):
        root = self.get_root(root_name)
        row_tag = "\n{%% row '%s' %%}" % root.rkey
        return XlRange.to_tpl(root) + row_tag

    def get_pos(self, wtsheet, root_name='main', nocache=False):
        root = self.get_root(root_name)
        pos_parent = self.pos_cls(wtsheet, root, nocache)
        for child in root._children_split:
            p = child.get_pos(wtsheet, pos_parent)
        return pos_parent


class HorizRange(XlRange):
    pos_cls = HRangePos

class VertRange(XlRange):
    pos_cls = VRangePos

class CellTag():

    def __init__(self, cell_tag=dict()):
        self.beforerow = ''
        self.beforecell = ''
        self.aftercell = ''
        if cell_tag:
            self.__dict__.update(cell_tag)

class RangeTag():

    def __init__(self, range_tag=dict()):
        self.beforerange = ''
        self.afterrange = ''
        if range_tag:
            self.__dict__.update(range_tag)

    def __str__(self):
        return self.beforerange + self.afterrange

    def __repr__(self):
        return self.beforerange + self.afterrange

    def merge(self, other):
        if isinstance(other, RangeTag):
            self.beforerange = other.beforerange + self.beforerange
            self.afterrange = other.afterrange + self.afterrange