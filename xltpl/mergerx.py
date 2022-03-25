# -*- coding: utf-8 -*-

from copy import copy, deepcopy
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
from .merger import MergeMixin


class MergerMixin():

    @property
    def to_merge(self):
        return bool(self._merge_list)

class CellMerge(MergeMixin):

    def __init__(self, cell_range, merger):
        self.merger = merger
        self.set_range()
        self._first_row = cell_range.min_row
        self._last_row = cell_range.max_row
        self._first_col = cell_range.min_col
        self._last_col = cell_range.max_col

    def new_range(self):
        if self.start_wtrowx==self.end_wtrowx and self.start_wtcolx==self.end_wtcolx:
            return
        range = CellRange(None, self.start_wtcolx, self.start_wtrowx, self.end_wtcolx, self.end_wtrowx)
        self.merger.add_new_range(range)

class CellMerger(MergerMixin):

    def __init__(self, sheet):
        self.range_list = []
        self._merge_list = []
        self.get_merge_list(sheet)

    def get_merge_list(self, sheet):
        for range in sheet.merged_cells:
            _merge = CellMerge(range, self)
            self._merge_list.append(_merge)

    def add_new_range(self, range):
        self.range_list.append(range)

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        for _merge in self._merge_list:
            is_in_range = _merge.merge_cell(rdrowx, rdcolx, wtrowx, wtcolx)
            if is_in_range:
                break

    def collect_range(self, wtsheet):
        for _merge in self._merge_list:
            _merge.collect_range()
        for range in self.range_list:
            wtsheet.merged_cells.add(range)
        self.range_list.clear()

class DataValidation(MergeMixin):

    def __init__(self, cell_range, merger, dv_key):
        self.dv_key = dv_key
        self.merger = merger
        self.set_range()
        self._first_row = cell_range.min_row
        self._last_row = cell_range.max_row
        self._first_col = cell_range.min_col
        self._last_col = cell_range.max_col

    def new_range(self):
        if self.start_wtrowx==-1:
            return
        range = CellRange(None, self.start_wtcolx, self.start_wtrowx, self.end_wtcolx, self.end_wtrowx)
        self.merger.add_new_range(self.dv_key, range)

class DvMerger(MergerMixin):

    def __init__(self, sheet):
        self.dv_map = {}
        self.dv_copy_map = {}
        self._merge_list = []
        self.get_merge_list(sheet)

    def get_merge_list(self, rdsheet):
        for index,dv in enumerate(rdsheet.data_validations.dataValidation):
            self.dv_map[index] = dv
            for crange in dv.ranges:
                _merge = DataValidation(crange, self, index)
                self._merge_list.append(_merge)

    def add_new_range(self, dv_key, range):
        dv_copy = self.dv_copy_map.get(dv_key)
        if not dv_copy:
            dv_copy = copy(self.dv_map[dv_key])
            dv_copy.ranges = MultiCellRange()
            self.dv_copy_map[dv_key] = dv_copy
        dv_copy.ranges.add(range)

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        for _merge in self._merge_list:
            is_in_range = _merge.merge_cell(rdrowx, rdcolx, wtrowx, wtcolx)
            if is_in_range:
                break

    def collect_range(self, wtsheet):
        for _merge in self._merge_list:
            _merge.collect_range()
        for key, dv in self.dv_copy_map.items():
            wtsheet.data_validations.append(dv)
        self.dv_copy_map.clear()

from collections import defaultdict
class ImageMerge(MergeMixin):

    def __init__(self, image, merger, image_count_dict):
        self.merger = merger
        self.image = image
        self.set_range()
        self.image_copy_map = {}
        self.image_ref_map = {}
        _from = image.anchor._from
        _to = image.anchor.to
        self._first_row = rlo = _from.row + 1
        self._first_col = clo = _from.col + 1
        self._last_row = rhi = _to.row + 1
        self._last_col = chi = _to.col + 1
        _top_left = (rlo, clo)
        count = image_count_dict[_top_left]
        image_count_dict[_top_left] += 1
        self.image_key = (rlo, clo, count)

    def new_range(self):
        if self.start_wtrowx==-1:
            return
        image = deepcopy(self.image)
        _from = image.anchor._from
        _to = image.anchor.to
        _from.row = self.start_wtrowx - 1
        _from.col = self.start_wtcolx - 1
        _to.row = self.end_wtrowx - 1
        _to.col = self.end_wtcolx - 1
        self.image_copy_map[(self.start_wtrowx, self.start_wtcolx)] = image

    def set_image_ref(self, image_ref):
        if image_ref.image:
            self.image_ref_map[image_ref.wt_top_left] = image_ref.image

    def collect_range(self):
        self.new_range()
        self.set_range()
        for key,image in self.image_copy_map.items():
            ref = self.image_ref_map.get(key)
            if ref:
                image.ref = ref
            self.merger.add_image(image)
        self.image_copy_map.clear()
        self.image_ref_map.clear()

class ImageMerger(MergerMixin):

    def __init__(self, sheet):
        self.images = []
        self._merge_map = {}
        self._merge_list = []
        self.max_row = 0
        self.max_col = 0
        self.get_merge_list(sheet)

    def get_merge_list(self, rdsheet):
        image_count_dict = defaultdict(int)
        for image in rdsheet._images:
            _merge = ImageMerge(image, self, image_count_dict)
            self._merge_map[_merge.image_key] = _merge
            self._merge_list.append(_merge)
            self.max_row = max(self.max_row, _merge._last_row)
            self.max_col = max(self.max_col, _merge._last_col)

    def add_image(self, image):
        self.images.append(image)

    def set_image_ref(self, image_ref):
        _merge = self._merge_map.get(image_ref.image_key)
        if not _merge:
            return
        _merge.set_image_ref(image_ref)

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        for _merge in self._merge_list:
            _merge.merge_cell(rdrowx, rdcolx, wtrowx, wtcolx)

    def collect_range(self, wtsheet):
        for _merge in self._merge_list:
            _merge.collect_range()
        wtsheet._images = self.images
        self.images = []

class AutoFilter(MergeMixin):

    def __init__(self, rdsheet):
        if not rdsheet.auto_filter.ref:
            self.to_merge = False
            return
        self.to_merge = True
        self.auto_filter = rdsheet.auto_filter
        self.set_range()
        cell_range = CellRange(rdsheet.auto_filter.ref)
        self._first_row = cell_range.min_row
        self._last_row = cell_range.max_row
        self._first_col = cell_range.min_col
        self._last_col = cell_range.max_col
        self.first_af = None

    def new_range(self):
        if self.start_wtrowx==-1:
            return
        if not self.first_af:
            self.first_af = CellRange(None, self.start_wtcolx, self.start_wtrowx,
                                      self.end_wtcolx, self.end_wtrowx)

    def collect_range(self, wtsheet):
        self.new_range()
        self.set_range()
        if wtsheet.auto_filter.ref:
            self.first_af = None
            return
        if self.first_af:
            wtsheet.auto_filter = copy(self.auto_filter)
            wtsheet.auto_filter.ref = self.first_af.coord
            self.first_af = None


class Merger:

    def __init__(self, rdsheet):
        cell_merger = CellMerger(rdsheet)
        dv_merger = DvMerger(rdsheet)
        self.image_merger = image_merger = ImageMerger(rdsheet)
        auto_filter = AutoFilter(rdsheet)
        _merger_list = [cell_merger, dv_merger, image_merger, auto_filter]
        self.merger_list = []
        for merger in _merger_list:
            if merger.to_merge:
                self.merger_list.append(merger)

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        for merger in self.merger_list:
            merger.merge_cell(rdrowx, rdcolx, wtrowx, wtcolx)

    def collect_range(self, wtsheet):
        for merger in self.merger_list:
            merger.collect_range(wtsheet)

    def set_image_ref(self, image_ref):
        self.image_merger.set_image_ref(image_ref)

