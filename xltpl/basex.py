# -*- coding: utf-8 -*-

import copy
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from .cellcontext import CellContextX

class SheetBase():

    def copy_sheet_settings(self):
        self.wtsheet.sheet_format = copy.copy(self.rdsheet.sheet_format)
        self.wtsheet.sheet_properties = copy.copy(self.rdsheet.sheet_properties)
        # copy print settings
        self.wtsheet.page_setup = copy.copy(self.rdsheet.page_setup)
        self.wtsheet.print_options = copy.copy(self.rdsheet.print_options)
        self.wtsheet._print_rows = copy.copy(self.rdsheet._print_rows)
        self.wtsheet._print_cols = copy.copy(self.rdsheet._print_cols)
        self.wtsheet._print_area = copy.copy(self.rdsheet._print_area)
        self.wtsheet.page_margins = copy.copy(self.rdsheet.page_margins)
        self.wtsheet.protection = copy.copy(self.rdsheet.protection)
        self.wtsheet.HeaderFooter = copy.copy(self.rdsheet.HeaderFooter)
        self.wtsheet.views = copy.copy(self.rdsheet.views)
        self.wtsheet._images = copy.copy(self.rdsheet._images)

    def copy_row_dimension(self, rdrowx, wtrowx):
        if wtrowx in self.wtrows:
            return
        dim = self.rdsheet.row_dimensions.get(rdrowx)
        if dim:
            self.wtsheet.row_dimensions[wtrowx] = copy.copy(dim)
            self.wtsheet.row_dimensions[wtrowx].worksheet = self.wtsheet
            self.wtrows.add(wtrowx)

    def copy_col_dimension(self, rdcolx, wtcolx):
        if wtcolx in self.wtcols:
            return
        rdkey = get_column_letter(rdcolx)
        rddim = self.rdsheet.column_dimensions.get(rdkey)
        if not rddim:
            return
        wtdim = copy.copy(rddim)
        if rdcolx != wtcolx:
            wtkey = get_column_letter(wtcolx)
            wtdim.index = wtkey
            d = wtcolx - rdcolx
            wtdim.min += d
            wtdim.max += d
        else:
            wtkey = rdkey
        self.wtsheet.column_dimensions[wtkey] = wtdim
        self.wtsheet.column_dimensions[wtkey].worksheet = self.wtsheet
        self.wtcols.add(wtcolx)

    def _cell(self, source_cell, rdrowx, rdcolx, wtrowx, wtcolx, value=None, data_type=None):
        target_cell = self.wtsheet.cell(column=wtcolx, row=wtrowx)
        if value is None:
            target_cell.value = source_cell._value
            target_cell.data_type = source_cell.data_type
        elif isinstance(value, STRING_TYPES) and value.startswith('='):
            target_cell.value = value
        elif data_type:
            target_cell._value = value
            target_cell.data_type = data_type
        else:
            #value, data_type = get_type(value)
            target_cell.value = value
            #target_cell.data_type = data_type
        if source_cell.has_style:
            target_cell._style = copy.copy(source_cell._style)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy.copy(source_cell.hyperlink)
        #if source_cell.comment:
        #    target_cell.comment = copy.copy(source_cell.comment)
        return target_cell

    def cell(self, source_cell, rdrowx, rdcolx, wtrowx, wtcolx, value=None, data_type=None):
        self.copy_row_dimension(rdrowx, wtrowx)
        self.copy_col_dimension(rdcolx, wtcolx)
        return self._cell(source_cell, rdrowx, rdcolx, wtrowx, wtcolx, value, data_type)

    def get_cell_context(self, cell_node, rv, cty):
        return CellContextX(self, cell_node, rv, cty)


class BookBase():

    def get_font(self, fontId):
        ifont = self.font_map.get(fontId)
        if ifont:
            return ifont
        else:
            font = self.workbook._fonts[fontId]
            ifont = InlineFont()
            ifont.rFont = font.name
            ifont.charset = font.charset
            ifont.family = font.family
            ifont.b = font.b
            ifont.i = font.i
            ifont.strike = font.strike
            ifont.outline = font.outline
            ifont.shadow = font.shadow
            ifont.condense = font.condense
            ifont.extend = font.extend
            ifont.color = font.color
            ifont.sz = font.sz
            ifont.u = font.u
            ifont.vertAlign = font.vertAlign
            ifont.scheme = font.scheme
            self.font_map[fontId] = ifont
            return ifont
