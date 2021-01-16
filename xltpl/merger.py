# -*- coding: utf-8 -*-

from copy import copy, deepcopy
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange

class Merge():

    def __init__(self, rdsheet, wtsheet):
        self.rdsheet = rdsheet
        self.wtsheet = wtsheet
        self.prepare()

    @classmethod
    def get_sheet_maps(cls, rdsheet):
        pass

    def prepare(self):
        pass

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        rdcoords2d = (rdrowx, rdcolx)
        if rdcoords2d in self._top_left_map:
            if self._merge_ranges.get(rdcoords2d):
                crange = self._merge_ranges.get(rdcoords2d)
                self.add_cell_range(rdcoords2d, crange)
            self._merge_ranges[rdcoords2d] = (wtrowx, wtrowx, wtcolx, wtcolx)
        else:
            _top_left = self._already_set.get(rdcoords2d)
            if _top_left:
                rlo, rhi, clo, chi = self._merge_ranges.get(_top_left)
                self._merge_ranges[_top_left] = (rlo, max(rhi, wtrowx), clo, max(chi, wtcolx))

    def add_cell_range(self, rdcoords2d, crange):
        pass

    def merge_mcell(self, rdrowx, rdcolx, wtrowx, wtcolx, wt_top):
        rdcoords2d = (rdrowx, rdcolx)
        if rdcoords2d in self._top_left_map:
            rlo, rhi, clo, chi = self._merge_ranges.get(rdcoords2d)
            self._merge_ranges[rdcoords2d] = (rlo, max(rhi, wtrowx), clo, max(chi, wtcolx))
        else:
            _top_left = self._already_set.get(rdcoords2d)
            if _top_left:
                rlo, rhi, clo, chi = self._merge_ranges.get(_top_left)
                self._merge_ranges[_top_left] = (rlo, max(rhi, wtrowx), clo, max(chi, wtcolx))
            else:
                self.merge_single_cell(wtrowx, wtcolx, wt_top, rdcoords2d)

    def merge_single_cell(self, wtrowx, wtcolx, wt_top, rdcoords2d):
        pass

    def finish(self):
        pass

class MergeCell(Merge):

    @classmethod
    def get_sheet_maps(cls, rdsheet):
        _map = {}
        _nfa = {}
        for crange in rdsheet.merged_cells.ranges:
            rlo, rhi, clo, chi = crange.min_row, crange.max_row, crange.min_col, crange.max_col
            _map[(rlo, clo)] = (rlo, rhi, clo, chi)
            for rowx in range(rlo, rhi + 1):
                for colx in range(clo, chi + 1):
                    _nfa[(rowx, colx)] = (rlo, clo)
        rdsheet.mc = (_map, _nfa)

    def prepare(self):
        self._merge_ranges = {}
        self._top_left_map, self._already_set = self.rdsheet.mc

    def add_cell_range(self, rdcoords2d, crange):
        rlo, rhi, clo, chi = crange
        cr = CellRange(min_row=rlo, max_row=rhi, min_col=clo, max_col=chi)
        self.wtsheet.merged_cells.add(cr)

    def merge_single_cell(self, wtrowx, wtcolx, wt_top, rdcoords2d):
        key = (wt_top, rdcoords2d)
        single_cell_cr = self._merge_ranges.get(key)
        if single_cell_cr:
            rlo, rhi, clo, chi = single_cell_cr
        else:
            rlo, rhi, clo, chi = wt_top, wtrowx, wtcolx, wtcolx
        self._merge_ranges[key] = (rlo, max(rhi, wtrowx), clo, max(chi, wtcolx))

    def finish(self):
        for key, crange in self._merge_ranges.items():
            self.add_cell_range(key, crange)
        self._merge_ranges.clear()

class DataValidation(Merge):

    @classmethod
    def get_sheet_maps(cls, rdsheet):
        _map = {}
        _nfa = {}
        _orig_map = {}
        for dv in rdsheet.data_validations.dataValidation:
            for crange in dv.ranges:
                rlo, rhi, clo, chi = crange.min_row, crange.max_row, crange.min_col, crange.max_col
                _map[(rlo, clo)] = (rlo, rhi, clo, chi)
                _orig_map[(rlo, clo)] = dv
                for rowx in range(rlo, rhi + 1):
                    for colx in range(clo, chi + 1):
                        _nfa[(rowx, colx)] = (rlo, clo)
        rdsheet.dv = _map, _nfa, _orig_map

    def prepare(self):
        self._merge_ranges = {}
        self.dv_copies = {}
        self._top_left_map, self._already_set, self._orig_map = self.rdsheet.dv

    def add_cell_range(self, rdcoords2d, crange):
        rlo, rhi, clo, chi = crange
        cr = CellRange(min_row=rlo, max_row=rhi, min_col=clo, max_col=chi)
        dv = self.dv_copies.get(rdcoords2d)
        if not dv:
            dv = copy(self._orig_map.get(rdcoords2d))
            dv.ranges = MultiCellRange()
            self.dv_copies[rdcoords2d] = dv
        dv.ranges.add(cr)

    def finish(self):
        for key,crange in self._merge_ranges.items():
            self.add_cell_range(key, crange)
        for key,dv in self.dv_copies.items():
            self.wtsheet.data_validations.append(dv)
        self.dv_copies.clear()
        self._merge_ranges.clear()

from collections import defaultdict

class Image(Merge):

    def __init__(self, image, image_list):
        self.image = image
        self.image_list = image_list
        self.prepare()

    @classmethod
    def get_sheet_maps(cls, rdsheet, image, image_dict):
        _nfa = {}
        _from = image.anchor._from
        _to = image.anchor.to
        rlo = _from.row + 1
        clo = _from.col + 1
        rhi = _to.row + 1
        chi = _to.col + 1
        _top_left = (rlo, clo)
        for rowx in range(rlo, rhi + 1):
            for colx in range(clo, chi + 1):
                _nfa[(rowx, colx)] = (rlo, clo)
        count = image_dict[_top_left]
        image_dict[_top_left] += 1
        image._top_left = _top_left
        image._bottom_right = (rhi, chi)
        image._already_set = _nfa
        image.image_key = (rlo, clo, count)

    def prepare(self):
        self._merge_range = None
        self._top_left = self.image._top_left
        self._already_set = self.image._already_set
        #self.image_key = self.image.image_key
        self.image_ref_map = {}
        self.image_copy_map = {}

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        rdcoords2d = (rdrowx, rdcolx)
        if rdcoords2d == self._top_left:
            if self._merge_range:
                self.add_image_copy(self._merge_range)
            self._merge_range = (wtrowx, wtrowx, wtcolx, wtcolx)
        else:
            _top_left = self._already_set.get(rdcoords2d)
            if _top_left and self._merge_range:
                rlo, rhi, clo, chi = self._merge_range
                self._merge_range = (rlo, max(rhi, wtrowx), clo, max(chi, wtcolx))

    def merge_mcell(self, rdrowx, rdcolx, wtrowx, wtcolx, wt_top):
        rdcoords2d = (rdrowx, rdcolx)
        if rdcoords2d == self._top_left:
            if self._merge_range:
                rlo, rhi, clo, chi = self._merge_range
                self._merge_range = (rlo, max(rhi, wtrowx), clo, max(chi, wtcolx))
        else:
            _top_left = self._already_set.get(rdcoords2d)
            if _top_left and self._merge_range:
                rlo, rhi, clo, chi = self._merge_range
                self._merge_range = (rlo, max(rhi, wtrowx), clo, max(chi, wtcolx))

    def set_image_ref(self, image_ref, key):
        if image_ref:
            self.image_ref_map[key] = image_ref

    def add_image_copy(self, crange):
        rlo, rhi, clo, chi = crange
        image = deepcopy(self.image)
        _from = image.anchor._from
        _to = image.anchor.to
        _from.row = rlo - 1
        _from.col = clo - 1
        _to.row = rhi - 1
        _to.col = chi - 1
        self.image_copy_map[(rlo,clo)] = image

    def finish(self):
        if self._merge_range:
            self.add_image_copy(self._merge_range)
        self._merge_range = None
        for key,image in self.image_copy_map.items():
            ref = self.image_ref_map.get(key)
            if ref:
                image.ref = ref
            self.image_list.add_image(image)
        self.image_copy_map.clear()
        self.image_ref_map.clear()

class ImageList(Merge):

    @classmethod
    def get_sheet_maps(cls, rdsheet):
        image_dict = defaultdict(int)
        for image in rdsheet._images:
            Image.get_sheet_maps(rdsheet, image, image_dict)

    def prepare(self):
        self.image_map = {}
        self.image_rhi = 0
        for image in self.rdsheet._images:
            self.image_map[image.image_key] = Image(image, self)
            rhi, chi = image._bottom_right
            self.image_rhi = max(self.image_rhi, rhi)
        self.images = []

    def set_image_ref(self, image_ref, image_key):
        rowx, colx, image_key = image_key
        image = self.image_map.get(image_key)

        if not image:
            return
        image.set_image_ref(image_ref, (rowx, colx))

    def add_image(self, image):
        self.images.append(image)

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        for img in self.image_map.values():
            img.merge_cell(rdrowx, rdcolx, wtrowx, wtcolx)

    def merge_mcell(self, rdrowx, rdcolx, wtrowx, wtcolx, wt_top):
        for img in self.image_map.values():
            img.merge_mcell(rdrowx, rdcolx, wtrowx, wtcolx, wt_top)

    def fix(self):
        wtsheet_rhi = self.wtsheet.max_row
        wtsheet_chi = self.wtsheet.max_column
        rdsheet_rhi = self.rdsheet.max_row
        for row in range(self.image_rhi - rdsheet_rhi):
            wtrowx = wtsheet_rhi + row + 1
            rdrowx = rdsheet_rhi + row + 1
            for colx in range(1, wtsheet_chi+1):
                self.merge_cell(rdrowx, colx, wtrowx, colx)

    def finish(self):
        self.fix()
        for img in self.image_map.values():
            img.finish()
        self.wtsheet._images = self.images

class Merger(object):

    def __init__(self, rdsheet, wtsheet):
        mc = MergeCell(rdsheet, wtsheet)
        dv = DataValidation(rdsheet, wtsheet)
        image_list = ImageList(rdsheet, wtsheet)
        self.image_list = image_list
        self.mergers = [mc, dv, image_list]

    @classmethod
    def get_sheet_maps(cls, rdsheet):
        MergeCell.get_sheet_maps(rdsheet)
        DataValidation.get_sheet_maps(rdsheet)
        ImageList.get_sheet_maps(rdsheet)

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        for merger in self.mergers:
            merger.merge_cell(rdrowx, rdcolx, wtrowx, wtcolx)

    def merge_mcell(self, rdrowx, rdcolx, wtrowx, wtcolx, wt_top):
        for merger in self.mergers:
            merger.merge_mcell(rdrowx, rdcolx, wtrowx, wtcolx, wt_top)

    def set_image_ref(self, image_ref, image_key):
        self.image_list.set_image_ref(image_ref, image_key)

    def finish(self):
        for merger in self.mergers:
            merger.finish()