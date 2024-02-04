# -*- coding: utf-8 -*-

from .patchx import *
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText
from .basex import SheetBase, BookBase
from .writermixin import SheetMixin, BookMixin, Box
from .utils import tag_test, parse_cell_tag
from .xlnode import Tree, Row, Cell, EmptyCell, Node, create_cell
from .jinja import JinjaEnvx
from .nodemap import NodeMap
from .sheetresource import SheetResourceMap
from .richtexthandler import rich_handlerx
from .mergerx import Merger
from .config import config
from .celltag import CellTag
from .image import img_cache

class SheetWriter(SheetBase, SheetMixin):

    def __init__(self, bookwriter, sheet_resource, sheet_name):
        self.workbook = bookwriter.workbook
        self.merger = sheet_resource.merger
        self.rdsheet = sheet_resource.rdsheet
        self.wtsheet = self.workbook.create_sheet(title=sheet_name)
        self.copy_sheet_settings()
        self.wtrows = set()
        self.wtcols = set()
        self.box = Box(0, 0)


class BookWriter(BookBase, BookMixin):
    sheet_writer_cls = SheetWriter

    def __init__(self, fname, debug=False):
        config.debug = debug
        self.load(fname)

    def load(self, fname):
        self.workbook = load_workbook(fname, rich_text=True)
        self.font_map = {}
        self.node_map = NodeMap()
        self.jinja_env = JinjaEnvx(self.node_map)
        self.merger_cls = Merger
        self.sheet_writer_map = {}
        self.sheet_resource_map = SheetResourceMap(self, self.jinja_env)
        for index,rdsheet in enumerate(self.workbook.worksheets):
            self.sheet_resource_map.add(rdsheet, rdsheet.title, index)
            self.workbook.remove(rdsheet)

    def build(self, sheet, index, merger):
        tree = Tree(index, self.node_map)
        max_row = max(sheet.max_row, merger.image_merger.max_row)
        max_col = max(sheet.max_column, merger.image_merger.max_col)
        for rowx in range(1, max_row + 1):
            row_node = Row(rowx)
            tree.add_child(row_node)
            for colx in range(1, max_col + 1):
                sheet_cell = sheet._cells.get((rowx, colx))
                if not sheet_cell:
                    cell_node = EmptyCell(rowx, colx)
                    tree.add_child(cell_node)
                    continue
                cell_tag_map = None
                if sheet_cell.comment:
                    comment = sheet_cell.comment.text
                    if tag_test(comment):
                        _,cell_tag_map = parse_cell_tag(comment)
                value = sheet_cell._value
                data_type = sheet_cell.data_type
                if data_type == 's':
                    rich_text = None
                    if isinstance(value, CellRichText):
                        #print(value)
                        rich_text = value
                        value = str(rich_text)
                    if not tag_test(value):
                        if rich_text:
                            cell_node = Cell(sheet_cell, rowx, colx, rich_text, data_type)
                        else:
                            cell_node = Cell(sheet_cell, rowx, colx, value, data_type)
                    else:
                        font = self.get_font(sheet_cell._style.fontId)
                        cell_node = create_cell(sheet_cell, rowx, colx, value, rich_text, data_type, font, rich_handlerx)
                else:
                    cell_node = Cell(sheet_cell, rowx, colx, value, data_type)
                if cell_tag_map:
                    cell_tag = CellTag(cell_tag_map)
                    cell_node.extend_cell_tag(cell_tag)
                    if colx==1:
                        row_node.cell_tag = cell_tag
                tree.add_child(cell_node)
        tree.add_child(Node())#
        return tree

    def cleanup_defined_names(self):
        self.workbook.custom_doc_props = ()
        # Custom Document Properties cause invalid file error
        sheet_cnt = len(self.workbook.worksheets)
        valid_names = {}
        for k, v in self.workbook.defined_names.items():
            if v.localSheetId:
                if int(v.localSheetId) < sheet_cnt:
                    valid_names[k] = v
            else:
                valid_names[k] = v
        self.workbook.defined_names = valid_names

    def save(self, fname):
        if not self.workbook.active:
            self.workbook.active = 0
        self.cleanup_defined_names()
        self.workbook.save(fname)
        img_cache.clear()
        for sheet in self.workbook.worksheets:
            self.workbook.remove(sheet)
        self.sheet_writer_map.clear()
